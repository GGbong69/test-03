#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 동전 표를 게임 데이터 테이블 꼴의 엑셀로 낸다.
#
#     python scripts/tools/make_item_xlsx.py
#
# 표 위에 스키마 다섯 줄(Comment/Min/Max/Scale/Type)을 얹는다. 열 뜻을
# 표 밖 문서에 두면 표와 갈리므로 **표 안에** 적는다. 뜻 문구는 손으로
# 짓지 않고 data.gd 가 화면에 찍는 문장을 그대로 옮긴다 — 그래야 표와
# 게임이 같은 말을 한다.
import csv
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except AttributeError:
        pass

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SRC = os.path.join(ROOT, "data", "items.csv")
DUMP = os.path.join(ROOT, "shots", "items_all.csv")   # 고닷이 찍어 둔 문구
OUT = os.path.join(ROOT, "shots", "표", "Item.xlsx")

HEAD_BG = "FF217346"
OFF_BG = "FFF2F2F2"
THIN = Side(style="thin", color="FFBFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
F8 = Font(name="맑은 고딕", size=8)
F10 = Font(name="맑은 고딕", size=10)

HEAD_ROW = 8
DATA_ROW = 9
INLINE_MAX = 4      # 이보다 값이 많은 이넘은 Enum 시트로 보낸다

# (열이름, 자료형, Min, Max, Scale, 설명, 이넘키)
COLS = [
    ("id", "INT", 110001, 149999, "-", "고유 번호. 1 + 희귀도 + 일련번호", None),
    ("name", "STRING", "-", "-", "-", "화면에 뜨는 이름", None),
    ("rarity", "ENUM", "-", "-", "-", "등급. id 둘째 자리와 같다", "rarity"),
    ("cost", "INT", 1, 99, "-", "상점 값. 판매가는 절반(내림·최소 2)", None),
    ("cond", "ENUM", "-", "-", "-",
     "언제 터지나. 비면 매 발. sec·col 은 뒤에 칸 번호를 붙인다", "cond"),
    ("kind", "ENUM", "-", "-", "-", "무엇을 주나. value 와 한 쌍", "kind"),
    ("value", "INT", 0, 999, "-", "kind 의 크기", None),
    ("kind_2", "ENUM", "-", "-", "-", "덧붙는 둘째 효과", "kind_2"),
    ("value_2", "INT", 0, 999, "-", "kind_2 의 크기", None),
    ("per", "ENUM", "-", "-", "-", "무엇의 개수만큼 곱하나", "per"),
    ("gold", "ENUM", "-", "-", "-", "점수와 별개로 주는 골드", "gold"),
    ("gv", "INT", 0, 99, "-", "gold 가 주는 골드", None),
    ("grow", "ENUM", "-", "-", "-", "값이 판을 거치며 변한다", "grow"),
    ("gstep", "INT", 0, 999, "-", "grow 의 한 번 폭", None),
    ("boom", "ENUM", "-", "-", "-", "스스로 부서질 확률", "boom"),
    ("dadd", "INT", -9, 9, "-", "판 시작 다트 증감", None),
    ("side", "ENUM", "-", "-", "-", "곁다리 효과", "side"),
    ("aim", "ENUM", "-", "-", "-",
     "조준 방식을 바꾼다. 있으면 cond·kind 는 빈다", "aim"),
    ("enabled", "ENUM", 0, 1, "-", "0 이면 안 나온다", "enabled"),
    ("_note", "STRING", "-", "-", "-", "왜 이 값인지. 게임이 안 읽는다", None),
]

# 표에 없는 열 → 원본 CSV 의 어느 칸에서 가져오나
FROM = {"kind_2": "k2", "value_2": "v2"}

WIDE = {"id": 9, "name": 16, "rarity": 11, "cond": 13, "kind": 12,
        "kind_2": 10, "value_2": 10, "gold": 10, "per": 12,
        "grow": 14, "boom": 11, "side": 13, "aim": 9,
        "enabled": 9, "_note": 58}
LEFT = ("name", "_note")

RAR = {"common": 1, "uncommon": 2, "rare": 3, "legendary": 4}

ENUM = {
    "rarity": [("common", "일반 — 등장 60%"), ("uncommon", "희귀 — 등장 35%"),
               ("rare", "레어 — 등장 5%"), ("legendary", "전설 — 아직 0장")],
    "kind": [("chip", "점수 +value"), ("mult", "배수 +value"),
             ("xmult", "배수 ×value"),
             ("mult_streak", "배수 +value (연속 명중 1회마다)"),
             ("mult_rand", "배수 +0~value 무작위"),
             ("save", "실패 1회 방지 (총점이 목표의 value% 이상일 때)")],
    "kind_2": [("mult", "배수 +value_2")],
    "per": [("darts_left", "남은 다트 1개당"), ("items", "보유 아이템 1개당"),
            ("gold", "보유 골드 1당"), ("gold5", "보유 골드 5당"),
            ("mag_hvy", "무거운 다트 1개당"),
            ("missing", "기본에서 줄어든 다트 1개당"),
            ("low", "판 최저 명중 숫자 1당"),
            ("zonehist", "이 영역의 런 누적 명중 1회당"),
            ("rackval", "다른 아이템 판매가 합계만큼"),
            ("empty", "빈 아이템 칸 수만큼 (최소 ×1)")],
    "gold": [("clear", "클리어 시 골드 +gv"), ("leg", "판마다 골드 +gv"),
             ("spare", "남은 다트 1개당 골드 +gv"),
             ("clean", "한 발도 안 빗나가면 골드 +gv"),
             ("blitz", "남은 다트가 기준 이상이면 골드 +gv"),
             ("broke", "정산 때 보유 골드가 기준 이하면 골드 +gv"),
             ("risk50", "더블·트리플·불 명중 시 절반 확률로 골드 +gv"),
             ("hit", "발동할 때마다 골드 +gv")],
    "grow": [("fire", "발동할 때마다 +gstep 누적"),
             ("hitmiss", "명중마다 +gstep · 빗나가면 −gstep"),
             ("tdec", "최댓값에서 시작 · 던질 때마다 −gstep"),
             ("rdec", "최댓값에서 시작 · 판마다 −gstep · 0이면 파괴")],
    "boom": [("r6", "판마다 1/6 확률로 파괴"),
             ("r1000", "판마다 0.1% 확률로 파괴")],
    "side": [("trackup25", "발동 4회 중 1회, 맞은 트랙 강화 +1")],
    "aim": [("ring", "원 — 원의 크기와 각도로 조준"),
            ("tilt", "빗각 — 조준선이 다트마다 다른 각도로 기움"),
            ("cross", "겹 — 가로세로가 함께 움직여 한 번에 잠김"),
            ("drift", "흔들 — 조준점이 커서 둘레를 떠돔"),
            ("place", "놓기 — 조준점을 원하는 자리에 직접 놓음"),
            ("pull", "당김 — 튕기는 속도와 방향으로 던짐"),
            ("kick", "반동 — 한 발이 작은 다트 5발, 쏠 때마다 조준이 밀림")],
    "enabled": [("1", "게임에 나온다"),
                ("0", "안 나온다 — 엔진이 아직 못 하는 효과")],
}


# 표 안에 적을 때만 쓰는 짧은 뜻. 칸 너비가 열 자 남짓이라 긴 말은 잘린다 —
# 잘려 보이느니 여기서 줄이고 온전한 뜻은 Enum 시트에 둔다.
SHORT = {
    "rarity": {"common": "일반", "uncommon": "희귀", "rare": "레어",
               "legendary": "전설"},
    "kind_2": {"mult": "배수 +value_2"},
    "grow": {"fire": "발동마다 누적", "hitmiss": "명중 +/ 빗나감 −",
             "tdec": "투척마다 감소", "rdec": "판마다 감소"},
    "boom": {"r6": "1/6 파괴", "r1000": "0.1% 파괴"},
    "side": {"trackup25": "4회 중 1회 트랙 +1"},
    "enabled": {"1": "쓴다", "0": "안 쓴다"},
}


def fold_cond(r):
    """sec·col 은 칸 번호를 조건 안에 접는다. 게임이 읽는 꼴과 같다 —
    data.gd:325 가 어차피 둘을 이어 "sec:4,10" 한 덩어리로 만든다."""
    c = r["cond"].strip()
    if c in ("sec", "col") and r["secs"].strip():
        return c + ":" + r["secs"].strip().replace(";", ",")
    return c


def head_cell(ws, row, col, text, bg=HEAD_BG):
    c = ws.cell(row, col, text)
    c.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFFFF")
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center")
    c.border = BOX
    return c


def build_item(ws, rows, dump, new):
    for i, lab in enumerate(("Comment", "Min", "Max", "Scale", "Type"), start=2):
        ws.cell(i, 1, lab).font = F8

    for j, (nm, typ, mn, mx, sc, desc, ek) in enumerate(COLS, start=2):
        cm = desc
        if ek:
            # 값이 넷을 넘으면 칸에 안 들어간다 — 잘려 보이느니 시트로 보낸다
            if len(ENUM[ek]) > INLINE_MAX:
                cm += "\n(%d종 — Enum 시트 참조)" % len(ENUM[ek])
            else:
                sh = SHORT.get(ek, {})
                cm += "\n" + "\n".join("%s : %s" % (k, sh.get(k, v))
                                       for k, v in ENUM[ek])
        for row, val in ((2, cm), (3, mn), (4, mx), (5, sc), (6, typ)):
            c = ws.cell(row, j, val)
            c.font = F8
            c.alignment = Alignment(
                wrap_text=(row == 2), vertical="top",
                horizontal=None if row == 2 else "center")
        h = head_cell(ws, HEAD_ROW, j, nm)
        ws.column_dimensions[h.column_letter].width = WIDE.get(nm, 10)

    # 읽기용 문구 — 게임은 안 읽는다. 초록 머리와 갈라 보이게 회색으로 둔다.
    dj = len(COLS) + 2
    h = head_cell(ws, HEAD_ROW, dj, "desc", "FF7F7F7F")
    ws.column_dimensions[h.column_letter].width = 42
    c = ws.cell(2, dj, "게임이 카드에 찍는 문장. 왼쪽 값에서 만들어진 것이라 "
                       "여기를 고쳐도 게임은 안 바뀐다")
    c.font = F8
    c.alignment = Alignment(wrap_text=True, vertical="top")

    for i, r in enumerate(sorted(rows, key=lambda x: new[x["id"]])):
        rr = DATA_ROW + i
        off = r["enabled"].strip() == "0"
        for j, (nm, typ, _mn, _mx, _sc, _d, _e) in enumerate(COLS, start=2):
            if nm == "id":
                v = new[r["id"]]
            elif nm == "cond":
                v = fold_cond(r)
            else:
                v = r[FROM.get(nm, nm)].strip()
            if typ in ("INT", "FLOAT") and str(v) != "":
                v = float(v) if typ == "FLOAT" else int(v)
            c = ws.cell(rr, j, v if str(v) != "" else None)
            c.font = F10
            c.border = BOX
            c.alignment = Alignment(horizontal="left" if nm in LEFT else "center")
            if off:
                c.fill = PatternFill("solid", start_color=OFF_BG)
        d = ws.cell(rr, dj, dump.get(r["id"], {}).get("효과", ""))
        d.font = Font(name="맑은 고딕", size=9, color="FF808080")
        d.border = BOX

    ws.row_dimensions[2].height = 96
    ws.freeze_panes = "C9"


def build_enum(ws, rows):
    for j, t in enumerate(("표", "열", "값", "뜻", "쓰는 줄"), start=1):
        head_cell(ws, 1, j, t)
    rr = 2
    for nm, _t, _mn, _mx, _sc, _d, ek in COLS:
        if not ek:
            continue
        used = {}
        for r in rows:
            v = fold_cond(r) if nm == "cond" else r.get(FROM.get(nm, nm), "").strip()
            if nm == "cond":
                v = v.split(":")[0]
            if v:
                used[v] = used.get(v, 0) + 1
        for k, meaning in ENUM[ek]:
            for j, val in enumerate(("Item", nm, k, meaning, used.get(k, 0)), start=1):
                c = ws.cell(rr, j, val)
                c.font = F10
                c.border = BOX
                c.alignment = Alignment(horizontal="left" if j == 4 else "center")
            rr += 1
    for col, w in (("A", 8), ("B", 10), ("C", 14), ("D", 54), ("E", 9)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    return rr - 2


def main():
    rows = list(csv.DictReader(open(SRC, encoding="utf-8-sig")))
    dump = {r["id"]: r for r in csv.DictReader(open(DUMP, encoding="utf-8-sig"))}

    # cond 의 뜻은 지어내지 않는다 — 게임이 찍는 조건 문구를 그대로 쓴다
    seen, cond = set(), []
    for r in rows:
        k = r["cond"].strip()
        if k and k not in seen:
            seen.add(k)
            cond.append((k, dump.get(r["id"], {}).get("조건", "")))
    ENUM["cond"] = sorted(cond)

    seq, new = {}, {}
    for r in rows:
        b = RAR[r["rarity"]]
        seq[b] = seq.get(b, 0) + 1
        new[r["id"]] = int("1%d%04d" % (b, seq[b]))

    wb = openpyxl.Workbook()
    build_item(wb.active, rows, dump, new)
    wb.active.title = "Item"
    n = build_enum(wb.create_sheet("Enum"), rows)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print("씀: %s" % OUT)
    print("  Item  %d행 · 열 %d + desc · 머리 5줄" % (len(rows), len(COLS)))
    print("  Enum  %d줄" % n)


if __name__ == "__main__":
    main()
